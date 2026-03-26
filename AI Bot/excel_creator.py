"""
Excel reports: signal history and ML performance (openpyxl).
Split from telegram_commands for clarity.
"""
import os
import json
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import config
import trade_manager
import path_utils


def parse_report_date(text: str) -> date | None:
    """
    Parse DD.MM.YYYY.
    Returns date or None for all data.
    Raises ValueError on bad format.
    """
    t = text.strip()
    if not t or t == "-":
        return None
    try:
        return datetime.strptime(t, "%d.%m.%Y").date()
    except ValueError:
        raise ValueError("Invalid date format. Use e.g. 10.02.2026")


def generate_excel_report(start_date: date | None = None):
    """
    Build Excel from signal_history.json.
    Returns (filepath, summary) or (None, error message).
    """
    history = trade_manager.load_history()
    signals = history.get("signals", [])
    if start_date:
        signals = [
            s for s in signals
            if s.get("start_time")
            and datetime.strptime(s["start_time"][:10], "%Y-%m-%d").date() >= start_date
        ]
    if not signals:
        return None, "No signals recorded yet."

    wb = Workbook()
    ws = wb.active
    ws.title = "Signal history"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ["ID", "Coin", "Signal", "Entry", "SL", "TP1", "TP2",
               "Start", "Status", "Close reason", "Close time", "P&L %",
               "Binance Order ID", "ML confidence"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    
    total_profit = 0
    closed_count = 0
    win_count = 0
    breakeven_count = 0

    for row_idx, signal in enumerate(signals, 2):
        ml_conf = signal.get("ml_confidence", "-")
        if isinstance(ml_conf, (int, float)):
            ml_conf_display = f"{ml_conf:.4f}"
        else:
            ml_conf_display = str(ml_conf)
        
        row_data = [
            signal.get("signal_id", ""),
            signal.get("symbol", ""),
            signal.get("signal", ""),
            signal.get("entry", ""),
            signal.get("sl", ""),
            signal.get("tp1", ""),
            signal.get("tp2", ""),
            signal.get("start_time", ""),
            signal.get("status", ""),
            signal.get("close_reason", "-"),
            signal.get("close_time", "-"),
            signal.get("profit_pct", "-"),
            signal.get("binance_order_id", "-"),
            ml_conf_display
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
        
        if signal.get("status") == "CLOSED" and signal.get("profit_pct") is not None:
            profit = signal.get("profit_pct", 0)
            total_profit += profit
            closed_count += 1
            if profit > 0:
                win_count += 1
            elif profit == 0:
                breakeven_count += 1

            if profit > 0:
                fill = green_fill
            elif profit == 0:
                fill = yellow_fill
            else:
                fill = red_fill
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = fill
    
    column_widths = [10, 12, 25, 12, 12, 12, 12, 20, 10, 15, 20, 12, 18, 14]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    summary_row = len(signals) + 3
    ws.cell(row=summary_row, column=1, value="SUMMARY").font = Font(bold=True, size=12)

    ws.cell(row=summary_row + 1, column=1, value="Total signals:")
    ws.cell(row=summary_row + 1, column=2, value=len(signals))

    ws.cell(row=summary_row + 2, column=1, value="Closed:")
    ws.cell(row=summary_row + 2, column=2, value=closed_count)

    ws.cell(row=summary_row + 3, column=1, value="Wins:")
    ws.cell(row=summary_row + 3, column=2, value=win_count)

    ws.cell(row=summary_row + 4, column=1, value="Breakeven:")
    ws.cell(row=summary_row + 4, column=2, value=breakeven_count)

    loss_count = closed_count - win_count - breakeven_count
    ws.cell(row=summary_row + 5, column=1, value="Losses:")
    ws.cell(row=summary_row + 5, column=2, value=loss_count)

    if closed_count > 0:
        win_rate = ((win_count + breakeven_count) / closed_count) * 100
        pure_win_rate = (win_count / closed_count) * 100
        ws.cell(row=summary_row + 6, column=1, value="Win rate (win+BE):")
        ws.cell(row=summary_row + 6, column=2, value=f"%{win_rate:.1f}")

        ws.cell(row=summary_row + 7, column=1, value="Pure win rate:")
        ws.cell(row=summary_row + 7, column=2, value=f"%{pure_win_rate:.1f}")

        ws.cell(row=summary_row + 8, column=1, value="Total P&L %:")
        total_cell = ws.cell(row=summary_row + 8, column=2, value=f"%{total_profit:.2f}")
        total_cell.font = Font(bold=True, color="006400" if total_profit > 0 else "8B0000")
    
    reports_dir = os.path.join(path_utils.get_base_dir(), "reports")
    if not os.path.exists(reports_dir):
        os.makedirs(reports_dir)

    filename = f"signal_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(reports_dir, filename)
    wb.save(filepath)
    
    summary_text = f"""Signal report ready.

Total signals: {len(signals)}
Closed: {closed_count}
Wins: {win_count}
Breakeven: {breakeven_count}
Losses: {closed_count - win_count - breakeven_count}
"""
    if closed_count > 0:
        summary_text += f"Win rate (win+BE): %{win_rate:.1f}\n"
        summary_text += f"Pure win rate: %{pure_win_rate:.1f}\n"
        summary_text += f"Total P&L: %{total_profit:.2f}"
    
    return filepath, summary_text


def generate_ai_excel_report(start_date: date | None = None):
    """
    Detailed AI report from ml_predictions.jsonl.
    KPIs: AI advantage, confidence buckets, underconfidence hints.
    start_date filters from this date (None = all).
    Returns (filepath, summary) or (None, error message).
    """
    jsonl_path = os.path.join(path_utils.get_base_dir(), "ml_predictions.jsonl")

    if not os.path.exists(jsonl_path):
        return None, "ML predictions file (ml_predictions.jsonl) not found."
        
    predictions = []
    try:
        with open(jsonl_path, 'r', encoding='utf-8') as f:
            for line in f:
                if line.strip():
                    predictions.append(json.loads(line))
    except Exception as e:
        return None, f"Read error: {e}"

    if start_date:
        filtered = []
        for p in predictions:
            ts = p.get("timestamp", "")
            if ts:
                try:
                    rec_date = datetime.fromisoformat(ts.replace("Z", "+00:00")).date()
                    if rec_date >= start_date:
                        filtered.append(p)
                except (ValueError, TypeError):
                    pass
        predictions = filtered

    if not predictions:
        return None, "No AI prediction data yet."

    completed_trades = [p for p in predictions if p.get('outcome') is not None or p.get('profit_pct') is not None]

    accepted_trades = [p for p in completed_trades if p.get('accepted')]
    rejected_trades = [p for p in completed_trades if not p.get('accepted')]
    
    # Helper: Win = profit_pct > 0, Breakeven = profit_pct == 0, Loss = profit_pct < 0
    def get_wins_breakeven_losses(trade_list):
        wins = [p for p in trade_list if (p.get('profit_pct') or 0) > 0]
        breakevens = [p for p in trade_list if p.get('profit_pct') is not None and p.get('profit_pct') == 0]
        losses = [p for p in trade_list if (p.get('profit_pct') or 0) < 0]
        return len(wins), len(breakevens), len(losses)

    acc_wins, acc_be, acc_losses = get_wins_breakeven_losses(accepted_trades)
    rej_wins, rej_be, rej_losses = get_wins_breakeven_losses(rejected_trades)
    
    acc_total = acc_wins + acc_be + acc_losses
    rej_total = rej_wins + rej_be + rej_losses
    
    acc_wr = ((acc_wins + acc_be) / acc_total * 100) if acc_total > 0 else 0
    rej_wr = ((rej_wins + rej_be) / rej_total * 100) if rej_total > 0 else 0

    acc_pure_wr = (acc_wins / acc_total * 100) if acc_total > 0 else 0
    rej_pure_wr = (rej_wins / rej_total * 100) if rej_total > 0 else 0

    ai_advantage = acc_wr - rej_wr

    _ = "HIGH" if (rej_wins > acc_wins and rej_total > 0) else "LOW"  # false-negative hint

    buckets = {
        "0.00-0.45": [],
        "0.45-0.50": [],
        "0.50-0.55": [],
        "0.55-0.60": [],
        "0.60+": []
    }
    
    for p in rejected_trades:
        conf = p.get('confidence', 0)
        if conf < 0.45:
            buckets["0.00-0.45"].append(p)
        elif 0.45 <= conf < 0.50:
            buckets["0.45-0.50"].append(p)
        elif 0.50 <= conf < 0.55:
            buckets["0.50-0.55"].append(p)
        elif 0.55 <= conf < 0.60:
            buckets["0.55-0.60"].append(p)
        else:
            buckets["0.60+"].append(p)
            
    bucket_stats = []
    underconfidence_flags = []
    
    for b_name, b_trades in buckets.items():
        count = len(b_trades)
        wins, be, losses = get_wins_breakeven_losses(b_trades)
        wr = ((wins + be) / count * 100) if count > 0 else 0
        
        total_pnl = sum([p.get('profit_pct') or 0 for p in b_trades])
        avg_pnl = (total_pnl / count) if count > 0 else 0
        
        bucket_stats.append({
            "bucket": b_name,
            "count": count,
            "wins": wins,
            "breakevens": be,
            "losses": losses,
            "win_rate": wr,
            "avg_pnl": avg_pnl
        })
        
        if wr > 60 and avg_pnl > 0:
            underconfidence_flags.append(b_name)

    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal="center")

    def style_header(ws, headers):
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
            
    ws_summary = wb.active
    ws_summary.title = "Summary KPI"

    ws_summary.cell(row=1, column=1, value="AI PERFORMANCE").font = Font(bold=True, size=12)
    ws_summary.cell(row=2, column=1, value="Metric").font = Font(bold=True)
    ws_summary.cell(row=2, column=2, value="Value").font = Font(bold=True)

    summary_rows = [
        ("Completed trades (with outcome)", len(completed_trades)),
        ("Accepted", acc_total),
        ("Rejected", rej_total),
        ("Accepted win rate (win+BE)", f"%{acc_wr:.1f} ({acc_wins}W + {acc_be}BE - {acc_losses}L)"),
        ("Accepted pure win rate", f"%{acc_pure_wr:.1f}"),
        ("Rejected win rate (win+BE)", f"%{rej_wr:.1f} ({rej_wins}W + {rej_be}BE - {rej_losses}L)"),
        ("Rejected pure win rate", f"%{rej_pure_wr:.1f}"),
        ("AI advantage", f"%{ai_advantage:+.1f}"),
        ("Underconfidence bands", ", ".join(underconfidence_flags) if underconfidence_flags else "NO"),
    ]
    
    for i, (metric, val) in enumerate(summary_rows, 3):
        ws_summary.cell(row=i, column=1, value=metric)
        ws_summary.cell(row=i, column=2, value=val)
        
    bucket_start_row = len(summary_rows) + 5
    ws_summary.cell(row=bucket_start_row, column=1, value="REJECTED CONFIDENCE BUCKETS").font = Font(bold=True, size=12)

    bucket_headers = ["Bucket", "Count", "Win", "BE", "Loss", "Win rate %", "Avg PnL %"]
    for col, h in enumerate(bucket_headers, 1):
        cell = ws_summary.cell(row=bucket_start_row + 1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = center_align

    for i, stat in enumerate(bucket_stats, bucket_start_row + 2):
        row_vals = [stat['bucket'], stat['count'], stat['wins'], stat['breakevens'], stat['losses'], f"%{stat['win_rate']:.1f}", f"%{stat['avg_pnl']:.2f}"]
        for col, val in enumerate(row_vals, 1):
            cell = ws_summary.cell(row=i, column=col, value=val)
            cell.alignment = center_align
            cell.border = border
            
            if col == 6:
                if stat['win_rate'] > 55: cell.fill = green_fill
            if col == 7:
                if stat['avg_pnl'] > 0: cell.fill = green_fill
                elif stat['avg_pnl'] < 0: cell.fill = red_fill

    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 15

    ws_details = wb.create_sheet("Details")

    dt_headers = ["Date", "Symbol", "Type", "Confidence", "Threshold", "Status", "Outcome", "P&L %", "Close reason", "Reason"]
    style_header(ws_details, dt_headers)
    
    for row_idx, p in enumerate(predictions, 2):
        status = "ACCEPTED" if p.get('accepted') else "REJECTED"
        profit = p.get('profit_pct')
        
        row_data = [
            p.get('timestamp', '').split('T')[0],
            p.get('symbol', ''),
            p.get('signal_type', ''),
            f"{p.get('confidence', 0):.4f}",
            p.get('threshold', ''),
            status,
            p.get('outcome', '-'),
            f"%{profit:.2f}" if profit is not None else "-",
            p.get('close_reason', '-'),
            p.get('reason', '')
        ]
        
        for col, val in enumerate(row_data, 1):
            cell = ws_details.cell(row=row_idx, column=col, value=val)
            cell.alignment = center_align
            cell.border = border
            
            if col == 8 and profit is not None:
                if profit > 0:
                    cell.fill = green_fill
                elif profit == 0:
                    cell.fill = yellow_fill
                else:
                    cell.fill = red_fill
            
            if col == 6:
                cell.font = Font(bold=True, color="006400" if status == "ACCEPTED" else "8B0000")

    for i in range(1, 11):
        ws_details.column_dimensions[get_column_letter(i)].width = 12
    ws_details.column_dimensions['B'].width = 12
    ws_details.column_dimensions['J'].width = 30

    reports_dir = os.path.join(path_utils.get_base_dir(), "reports")
    if not os.path.exists(reports_dir):
        os.makedirs(reports_dir)
        
    filename = f"ai_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(reports_dir, filename)
    wb.save(filepath)
    
    calibration_needed = "YES" if underconfidence_flags else "NO"

    summary_text = f"""AI performance summary

AI_FILTER_STATUS:
• Accepted WR: %{acc_wr:.1f}
• Rejected WR: %{rej_wr:.1f}
• AI advantage: %{ai_advantage:+.1f}
• Underconfidence: {", ".join(underconfidence_flags) if underconfidence_flags else "NONE"}
• Calibration needed: {calibration_needed}

Open the Excel file for full detail.
"""

    if underconfidence_flags:
        summary_text += "\nACTION:\n"
        for band in underconfidence_flags:
            summary_text += f"- Consider SMALL SIZE trades for confidence {band}\n"
        summary_text += "- Or recalibrate model (Platt / Isotonic)"

    return filepath, summary_text
